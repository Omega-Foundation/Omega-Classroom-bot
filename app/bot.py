"""Telegram bot handlers and commands."""
from telegram import Update
from telegram.ext import Application, CommandHandler, ContextTypes
from sqlalchemy import and_, func, or_
from sqlalchemy.orm import joinedload
from app.database import User, Assignment, TrackedRepository, Submission, ClassroomAssignmentRecord, get_db, init_db
from app.github_client import GitHubClient
from datetime import datetime, timezone
from app.config import Config
from dateutil import parser as date_parser
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from typing import Tuple, List, Dict, Optional
from collections import Counter
import re
import json

class HomeworkTrackerBot:
    """Main bot class."""
    
    def __init__(self):
        self.teacher_password = Config.TEACHER_ACCESS_PASSWORD.strip() if Config.TEACHER_ACCESS_PASSWORD else ''
    
    def _normalize_assignment_slug(self, assignment: dict) -> str:
        """Return a normalized slug for an assignment."""
        if not isinstance(assignment, dict):
            return ''
        slug = assignment.get('slug')
        if isinstance(slug, str):
            slug = slug.strip()
            if slug:
                return slug
        title = assignment.get('title') or assignment.get('name') or ''
        if not isinstance(title, str):
            title = ''
        normalized = re.sub(r'[^a-z0-9]+', '-', title.lower()).strip('-')
        return normalized

    def _parse_datetime(self, value) -> Optional[datetime]:
        """Parse various datetime representations into naive UTC."""
        if isinstance(value, datetime):
            dt = value
        elif isinstance(value, str):
            try:
                dt = date_parser.parse(value)
            except Exception:
                return None
        else:
            return None
        if dt.tzinfo:
            dt = dt.astimezone(timezone.utc)
        return dt.replace(tzinfo=None)

    def _extract_student_identity(self, assignment: dict, acceptance: dict) -> Tuple[str, str, str]:
        """
        Derive display login, repository URL, and canonical login for a student's acceptance.
        
        Returns:
            (display_login, repository_url, canonical_login)
        """
        def pick_value(values):
            for value in values:
                if isinstance(value, str):
                    value = value.strip()
                    if value:
                        return value
            return ''

        student = acceptance.get('student') or {}
        primary_login = pick_value([
            student.get('github_username'),
            student.get('login'),
            student.get('github_login'),
            student.get('name'),
            student.get('display_name'),
            acceptance.get('github_username'),
            acceptance.get('login'),
            acceptance.get('github_login'),
            acceptance.get('student_login'),
            acceptance.get('student_name'),
        ])

        display_login = primary_login

        repo = acceptance.get('repository') or {}
        repo_url = pick_value([
            repo.get('html_url'),
            repo.get('url'),
        ])
        repo_full_name = pick_value([
            repo.get('full_name'),
            acceptance.get('repository_full_name'),
        ])
        repo_name = pick_value([repo.get('name')])

        if not repo_url and repo_full_name:
            repo_url = f"https://github.com/{repo_full_name}"

        repo_path = ''
        if repo_url:
            repo_path = repo_url.split('github.com/', 1)[-1]
        elif repo_full_name:
            repo_path = repo_full_name

        owner_candidate = ''
        repo_slug = ''
        if repo_path:
            parts = repo_path.split('/', 2)
            if parts:
                owner_candidate = parts[0].strip()
            if len(parts) >= 2:
                repo_slug = parts[1].strip()

        assignment_slug = self._normalize_assignment_slug(assignment)
        assignment_slug_lower = assignment_slug.lower() if assignment_slug else ''

        def trim_candidate(candidate: str) -> str:
            candidate = candidate or ''
            candidate_lower = candidate.lower()
            if assignment_slug_lower:
                for sep in ('-', '_'):
                    prefix = f"{assignment_slug_lower}{sep}"
                    if candidate_lower.startswith(prefix):
                        trimmed = candidate[len(prefix):]
                        if trimmed:
                            return trimmed
                if candidate_lower.startswith(assignment_slug_lower):
                    trimmed = candidate[len(assignment_slug_lower):].lstrip('-_')
                    if trimmed:
                        return trimmed
            for sep in ('-', '_'):
                if sep in candidate:
                    parts = candidate.split(sep, 1)
                    if parts[1]:
                        return parts[1]
            return candidate

        login_lower = display_login.lower() if display_login else ''
        if (not display_login) or (assignment_slug_lower and login_lower == assignment_slug_lower) or (owner_candidate and login_lower == owner_candidate.lower()):
            candidate = repo_slug or repo_name or owner_candidate or display_login
            if candidate:
                trimmed = trim_candidate(candidate)
                candidate = trimmed or candidate
            display_login = candidate or display_login

        if (not display_login or (assignment_slug_lower and display_login.lower() == assignment_slug_lower)) and owner_candidate:
            display_login = owner_candidate

        display_login = display_login or "Unknown"
        canonical_login = primary_login or owner_candidate or repo_slug or repo_name or ""
        canonical_login = canonical_login.strip()
        return display_login, repo_url, canonical_login

    def _sync_assignment_record(
        self,
        db,
        teacher: User,
        classroom_id: Optional[int],
        classroom_name: Optional[str],
        assignment: Dict,
        accepted: List[Dict]
    ):
        """Persist classroom assignment info and link students based on GitHub usernames."""
        if not assignment or not teacher:
            return

        assignment_name = assignment.get('title') or assignment.get('name') or 'Classroom Assignment'
        assignment_desc = assignment.get('description')
        assignment_id = assignment.get('id')
        assignment_slug = self._normalize_assignment_slug(assignment)
        classroom_id_str = str(classroom_id) if classroom_id is not None else None
        assignment_id_str = str(assignment_id) if assignment_id is not None else None

        deadline_val = self._parse_datetime(assignment.get('deadline'))
        if deadline_val is None:
            deadline_val = datetime.utcnow()

        repo_url_candidates = [
            assignment.get('student_repository_url'),
            assignment.get('invitations_url'),
        ]
        starter_repo = assignment.get('starter_code_repository') or {}
        if isinstance(starter_repo, dict):
            repo_url_candidates.extend([
                starter_repo.get('html_url'),
                starter_repo.get('url'),
            ])
        assignment_repo_url = next((v for v in repo_url_candidates if isinstance(v, str) and v.strip()), '')
        assignment_repo_name = assignment_slug or assignment_name

        query = db.query(Assignment).filter(Assignment.user_id == teacher.id)
        if assignment_id_str:
            query = query.filter(Assignment.classroom_assignment_id == assignment_id_str)
        elif classroom_id_str:
            query = query.filter(Assignment.classroom_id == classroom_id_str, Assignment.name == assignment_name)
        existing_assignment = query.first()

        if not existing_assignment:
            existing_assignment = Assignment(
                name=assignment_name,
                description=assignment_desc,
                github_repo_name=assignment_repo_name or assignment_name,
                github_repo_url=assignment_repo_url,
                deadline=deadline_val,
                classroom_id=classroom_id_str,
                classroom_assignment_id=assignment_id_str,
                user_id=teacher.id,
            )
            db.add(existing_assignment)
            db.commit()
            db.refresh(existing_assignment)
        else:
            changed = False
            if assignment_desc and existing_assignment.description != assignment_desc:
                existing_assignment.description = assignment_desc
                changed = True
            if assignment_repo_url and existing_assignment.github_repo_url != assignment_repo_url:
                existing_assignment.github_repo_url = assignment_repo_url
                changed = True
            if assignment_repo_name and existing_assignment.github_repo_name != assignment_repo_name:
                existing_assignment.github_repo_name = assignment_repo_name
                changed = True
            if classroom_id_str and existing_assignment.classroom_id != classroom_id_str:
                existing_assignment.classroom_id = classroom_id_str
                changed = True
            if assignment_id_str and existing_assignment.classroom_assignment_id != assignment_id_str:
                existing_assignment.classroom_assignment_id = assignment_id_str
                changed = True
            if deadline_val and existing_assignment.deadline != deadline_val:
                existing_assignment.deadline = deadline_val
                changed = True
            if changed:
                db.commit()

        classroom_label = classroom_name or 'Classroom'

        for acceptance in accepted:
            display_login, repo_url, canonical_login = self._extract_student_identity(assignment, acceptance)
            canonical_lower = canonical_login.lower() if canonical_login else ''
            student_user = None
            if canonical_lower:
                student_user = db.query(User).filter(
                    func.lower(User.github_username) == canonical_lower
                ).first()
            if not student_user:
                continue

            submission = db.query(Submission).filter(
                Submission.assignment_id == existing_assignment.id,
                Submission.user_id == student_user.id
            ).first()

            submitted_flag = bool(acceptance.get('submitted'))
            submitted_at = self._parse_datetime(
                acceptance.get('submitted_at') or acceptance.get('updated_at') or acceptance.get('created_at')
            )

            if not submission:
                submission = Submission(
                    assignment_id=existing_assignment.id,
                    user_id=student_user.id,
                    github_repo_url=repo_url,
                    is_submitted=submitted_flag,
                    submitted_at=submitted_at,
                    created_at=datetime.utcnow(),
                )
                db.add(submission)
            else:
                changed = False
                if repo_url and submission.github_repo_url != repo_url:
                    submission.github_repo_url = repo_url
                    changed = True
                if submitted_flag != submission.is_submitted:
                    submission.is_submitted = submitted_flag
                    changed = True
                if submitted_at and submission.submitted_at != submitted_at:
                    submission.submitted_at = submitted_at
                    changed = True
                if changed:
                    submission.updated_at = datetime.utcnow()

        db.commit()

    def _store_classroom_records(
        self,
        db,
        teacher: User,
        records: List[Dict]
    ):
        """Persist classroom assignment snapshot records for a teacher."""
        if not teacher or not isinstance(records, list):
            return
        teacher_id = teacher.id
        if not teacher_id:
            return
        db.query(ClassroomAssignmentRecord).filter(
            ClassroomAssignmentRecord.teacher_user_id == teacher_id
        ).delete()
        for record in records:
            try:
                deadline = record.get('deadline')
                if isinstance(deadline, str):
                    deadline = self._parse_datetime(deadline)
                raw_payload = record.get('raw')
                if raw_payload is not None:
                    raw_json = json.dumps(raw_payload, default=str)
                else:
                    raw_json = None
                classroom_record = ClassroomAssignmentRecord(
                    teacher_user_id=teacher_id,
                    classroom_id=str(record.get('classroom_id') or ''),
                    classroom_name=record.get('classroom_name') or '',
                    assignment_id=str(record.get('assignment_id') or ''),
                    assignment_title=record.get('assignment_title') or '',
                    assignment_url=record.get('assignment_url') or '',
                    deadline=deadline,
                    student_login=record.get('student_login') or '',
                    student_display_login=record.get('student_display_login') or '',
                    student_repo_url=record.get('student_repo_url') or '',
                    submitted=record.get('submitted'),
                    passed=record.get('passed'),
                    grade=record.get('grade'),
                    commit_count=record.get('commit_count'),
                    raw_json=raw_json,
                )
                db.add(classroom_record)
            except Exception as exc:
                print(f"Failed to persist classroom record: {exc}")
        db.commit()

    def get_db(self):
        """Get database session."""
        db_gen = get_db()
        return next(db_gen)

    def _format_classroom_label(self, raw: Optional[str]) -> str:
        """Format classroom identifiers into a readable label."""
        if not raw or not isinstance(raw, str):
            return ""
        label = raw.replace('-', ' ').replace('_', ' ').strip()
        if not label:
            return ""
        return label.title()
    
    async def start(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /start command."""
        user = update.effective_user
        chat_id = update.effective_chat.id
        
        db = self.get_db()
        try:
            # Check if user exists
            db_user = db.query(User).filter(User.telegram_id == chat_id).first()
            
            if not db_user:
                # Create new user
                db_user = User(
                    telegram_id=chat_id,
                    username=user.username,
                    first_name=user.first_name,
                    last_name=user.last_name
                )
                db.add(db_user)
                db.commit()
                db.refresh(db_user)
            
            role = (db_user.role or 'student').lower()
            role_display = "Teacher" if role == 'teacher' else "Student"

            # Check if user has GitHub token
            if not db_user.github_token:
                welcome_message = [
                    f"👋 Welcome, {user.first_name}!",
                    "",
                    "I'm your Omega Classroom tracking bot.",
                    "",
                    "To get started, please provide your GitHub personal access token:",
                    "/register_token <your_github_token>",
                    "",
                    "You can create a token at: https://github.com/settings/tokens",
                    "Required permissions: repo, read:org",
                ]
            else:
                welcome_message = [
                    f"Welcome back, {user.first_name}! ({role_display})",
                    "",
                    "Available commands:",
                    "/assignments - List your assignments",
                    "/help - Show help",
                ]
                if role == 'teacher':
                    welcome_message.extend([
                        "/add_assignment - Add a new assignment",
                        "/delete_assignment - Delete an assignment",
                        "/add_note - Add a note to an assignment",
                        "/delete_note - Delete a note from an assignment",
                        "/classroom_assignments - Classroom overview",
                        "/export_assignments_excel - Export classroom data",
                        "/dump_submissions - Dump submissions table",
                        "/export_users_excel - Export user list",
                        "/ci_status - Show tracked CI status",
                    ])
            if not db_user.github_username:
                welcome_message.extend([
                    "",
                    "I don't know your GitHub username yet.",
                    "Please set it with: /set_github_username <github_login>",
                ])
            welcome_message.extend([
                "",
                "Need to switch roles? Use /set_role <student|teacher> [password]",
            ])
            
            await update.message.reply_text('\n'.join(welcome_message))
        finally:
            db.close()
    
    async def help_command(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /help command."""
        help_text = (
            "📚 Omega Classroom Bot Commands\n\n"
            "/start - Start the bot\n"
            "/register_token <token> - Register your GitHub personal access token\n"
            "/assignments - List all your assignments\n"
            "/add_ci_repo <repo> - Track GitHub Actions status for a repository\n"
            "/remove_ci_repo <repo> - Stop tracking repository CI status\n"
            "/ci_status - Show latest CI status for tracked repositories\n"
            "/set_my_notify_threshold <days> - Start notifications N days before deadline (you)\n"
            "/set_my_notify_period <value><m|h> - Reminder interval for you (e.g. 60m, 1h)\n"
            "/set_role <student|teacher> [password] - Switch between student and teacher roles\n"
            "/set_github_username <username> - Manually set your GitHub username\n"
            "\nTeacher-only commands:\n"
            "/add_assignment - Add a new assignment\n"
            "/delete_assignment - Delete an assignment\n"
        "/add_note <assignment_name> <text> - Add a note to an assignment\n"
        "/delete_note <assignment_name> - Delete a note from an assignment\n"
            "/classroom_assignments - View classroom assignments overview\n"
            "/export_assignments_excel - Export classroom data to Excel\n"
            "/dump_submissions - Dump submissions table\n"
            "/export_users_excel - Export user information to Excel\n"
            "/help - Show this help message\n"
        )
        await update.message.reply_text(help_text)
    
    async def register_token(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /register_token command."""
        chat_id = update.effective_chat.id
        
        if not context.args:
            await update.message.reply_text(
                "Please provide your GitHub personal access token:\n"
                "/register_token <your_github_token>\n\n"
                "You can create a token at: https://github.com/settings/tokens\n"
                "Required permissions: repo, read:org"
            )
            return
        
        github_token = context.args[0]
        
        # Validate token by trying to create a GitHub client
        try:
            test_client = GitHubClient(token=github_token)
            # Try to get user info to validate token
            test_user = test_client.github.get_user()
            github_username = test_user.login
        except Exception as e:
            await update.message.reply_text(
                f"❌ Invalid GitHub token. Please check your token and try again.\n"
                f"Error: {str(e)}\n\n"
                f"Create a token at: https://github.com/settings/tokens"
            )
            return
        
        db = self.get_db()
        try:
            # Update user
            db_user = db.query(User).filter(User.telegram_id == chat_id).first()
            if db_user:
                db_user.github_token = github_token
                if github_username:
                    db_user.github_username = github_username
                db.commit()
                display_username = github_username or db_user.github_username
                message_lines = [
                    "✅ GitHub token registered successfully!",
                    "",
                    f"GitHub username: {display_username or 'Not detected'}",
                    "Your token has been saved securely.",
                    "",
                    "Next steps:",
                    "/assignments - View your assignments",
                ]
                if (db_user.role or 'student').lower() == 'teacher':
                    message_lines.append("/add_assignment - Add a new assignment")
                if not github_username:
                    message_lines.extend([
                        "",
                        "I could not determine your GitHub username.",
                        "Please set it manually: /set_github_username <github_login>"
                    ])
                await update.message.reply_text('\n'.join(message_lines))
            else:
                await update.message.reply_text(
                    "Please use /start first to register."
                )
        finally:
            db.close()

    async def list_assignments(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /assignments command - get assignments from GitHub Classroom API."""
        chat_id = update.effective_chat.id

        db = self.get_db()
        try:
            db_user = db.query(User).filter(User.telegram_id == chat_id).first()

            if not db_user:
                await update.message.reply_text("Please use /start first.")
                return

            if not db_user.github_username:
                await update.message.reply_text(
                    "Please set your GitHub username so I can look up your assignments:\n"
                    "/set_github_username <github_login>"
                )
                return

            if not db_user.github_token:
                await update.message.reply_text(
                    "Please register your GitHub token first:\n"
                    "/register_token <your_github_token>"
                )
                return

            try:
                github_client = GitHubClient(token=db_user.github_token)
            except Exception as e:
                await update.message.reply_text(f"❌ Could not initialise GitHub client: {e}")
                return

            classroom_assignments: List[Dict] = []
            classroom_meta: Dict[str, Dict] = {}
            fetch_error: Optional[str] = None

            try:
                classroom_assignments = github_client.get_classroom_assignments(db_user.github_username)
            except Exception as e:
                fetch_error = str(e)
            else:
                for item in classroom_assignments:
                    assignment_name = (item.get('name') or item.get('title') or 'Classroom Assignment').strip()
                    classroom_id = item.get('classroom_id')
                    assignment_id = item.get('assignment_id')
                    assignment_id_str = str(assignment_id) if assignment_id is not None else None
                    classroom_id_str = str(classroom_id) if classroom_id is not None else None

                    if assignment_id_str:
                        classroom_meta[assignment_id_str] = item

                    deadline_dt = self._parse_datetime(item.get('deadline')) or datetime.utcnow()
                    repo_url = (item.get('url') or '').strip()
                    repo_name = github_client.parse_repo_url(repo_url) if repo_url else assignment_name

                    query = db.query(Assignment).filter(Assignment.user_id == db_user.id)
                    if assignment_id_str:
                        query = query.filter(Assignment.classroom_assignment_id == assignment_id_str)
                    else:
                        query = query.filter(Assignment.name == assignment_name)
                    assignment_db = query.first()

                    changed = False
                    if not assignment_db:
                        assignment_db = Assignment(
                            name=assignment_name,
                            description=item.get('description'),
                            github_repo_name=repo_name or assignment_name,
                            github_repo_url=repo_url,
                            deadline=deadline_dt,
                            classroom_id=classroom_id_str,
                            classroom_assignment_id=assignment_id_str,
                            user_id=db_user.id,
                        )
                        db.add(assignment_db)
                        db.flush()
                        changed = True
                    else:
                        if item.get('description') and assignment_db.description != item.get('description'):
                            assignment_db.description = item.get('description')
                            changed = True
                        if repo_url and assignment_db.github_repo_url != repo_url:
                            assignment_db.github_repo_url = repo_url
                            changed = True
                        if repo_name and assignment_db.github_repo_name != repo_name:
                            assignment_db.github_repo_name = repo_name
                            changed = True
                        if classroom_id_str and assignment_db.classroom_id != classroom_id_str:
                            assignment_db.classroom_id = classroom_id_str
                            changed = True
                        if assignment_id_str and assignment_db.classroom_assignment_id != assignment_id_str:
                            assignment_db.classroom_assignment_id = assignment_id_str
                            changed = True
                        if deadline_dt and assignment_db.deadline != deadline_dt:
                            assignment_db.deadline = deadline_dt
                            changed = True

                    submission = next((s for s in assignment_db.submissions or [] if s.user_id == db_user.id), None)
                    if not submission:
                        submission = Submission(
                            assignment_id=assignment_db.id,
                            user_id=db_user.id,
                            github_repo_url=repo_url or assignment_db.github_repo_url,
                        )
                        db.add(submission)
                        changed = True
                    else:
                        if repo_url and submission.github_repo_url != repo_url:
                            submission.github_repo_url = repo_url
                            submission.updated_at = datetime.utcnow()
                            changed = True

                    if changed:
                        db.commit()
                        db.refresh(assignment_db)

            assignments = (
                db.query(Assignment)
                .options(joinedload(Assignment.user), joinedload(Assignment.submissions))
                .filter(Assignment.user_id == db_user.id)
                .order_by(Assignment.deadline, Assignment.name)
                .all()
            )

            class_ids = {a.classroom_assignment_id for a in assignments if a.classroom_assignment_id}
            teacher_map: Dict[str, str] = {}
            classroom_teacher_map: Dict[str, str] = {}
            if class_ids:
                class_id_list = [cid for cid in class_ids if cid]
                if class_id_list:
                    teacher_assignments = (
                        db.query(Assignment)
                        .options(joinedload(Assignment.user))
                        .filter(
                            Assignment.classroom_assignment_id.in_(class_id_list),
                            Assignment.user_id != db_user.id
                        )
                        .all()
                    )
                    for teacher_assignment in teacher_assignments:
                        if teacher_assignment.classroom_assignment_id and teacher_assignment.user:
                            teacher_user = teacher_assignment.user
                            teacher_name = teacher_user.first_name or teacher_user.username or teacher_user.github_username or ""
                            if teacher_name:
                                teacher_map[str(teacher_assignment.classroom_assignment_id)] = teacher_name

                    classroom_records = (
                        db.query(ClassroomAssignmentRecord)
                        .options(joinedload(ClassroomAssignmentRecord.teacher))
                        .filter(ClassroomAssignmentRecord.assignment_id.in_(class_id_list))
                        .all()
                    )
                    for record in classroom_records:
                        record_key = record.assignment_id or ""
                        if not record_key:
                            continue
                        teacher_entity = record.teacher
                        teacher_name = ""
                        if teacher_entity:
                            teacher_name = teacher_entity.first_name or teacher_entity.username or teacher_entity.github_username or ""
                        if not teacher_name:
                            teacher_name = self._format_classroom_label(record.classroom_name)
                        if teacher_name:
                            classroom_teacher_map.setdefault(str(record_key), teacher_name)

            entries: List[Dict] = []
            seen_keys = set()
            now = datetime.utcnow()

            for assignment in assignments:
                key_identifier = assignment.classroom_assignment_id or assignment.id or assignment.name or ''
                key = str(key_identifier).lower()
                if key in seen_keys:
                    continue
                seen_keys.add(key)

                submission = next((s for s in assignment.submissions or [] if s.user_id == db_user.id), None)
                repo_url = ''
                if submission and submission.github_repo_url:
                    repo_url = submission.github_repo_url
                else:
                    repo_url = assignment.github_repo_url or assignment.github_repo_name or ''

                deadline = assignment.deadline if isinstance(assignment.deadline, datetime) else None
                status = "ℹ️"
                time_remaining = ""
                if deadline:
                    if deadline > now:
                        delta = deadline - now
                        status = "⏰ Active"
                        time_remaining = f" ({delta.days}d {delta.seconds // 3600}h remaining)"
                    else:
                        status = "✅ Past"

                submission_status = None
                submitted_at_str = None
                if submission:
                    if submission.is_submitted is not None:
                        submission_status = "Submitted" if submission.is_submitted else "In progress"
                    if submission.submitted_at and isinstance(submission.submitted_at, datetime):
                        submitted_at_str = submission.submitted_at.strftime('%Y-%m-%d %H:%M:%S UTC')

                teacher_name = ""
                if assignment.classroom_assignment_id:
                    teacher_name = teacher_map.get(str(assignment.classroom_assignment_id), "")
                    if not teacher_name:
                        teacher_name = classroom_teacher_map.get(str(assignment.classroom_assignment_id), "")
                    if not teacher_name:
                        meta = classroom_meta.get(str(assignment.classroom_assignment_id))
                        if meta:
                            teacher_name = self._format_classroom_label(meta.get('classroom_name'))
                elif assignment.user and assignment.user.id != db_user.id:
                    teacher_user = assignment.user
                    teacher_name = teacher_user.first_name or teacher_user.username or teacher_user.github_username or ""

                source_label = "Classroom" if assignment.classroom_assignment_id else "Saved"

                entries.append({
                    'name': assignment.name,
                    'deadline': deadline,
                    'repo_url': repo_url,
                    'teacher': teacher_name,
                    'status': status,
                    'time_remaining': time_remaining,
                    'submission_status': submission_status,
                    'submitted_at': submitted_at_str,
                    'source': source_label,
                    'note': (assignment.note or '').strip(),
                })

            if not entries:
                message_lines = [
                    "No assignments found.",
                    "You can add assignments with: /add_assignment"
                ]
                if fetch_error:
                    message_lines.append("")
                    message_lines.append(f"⚠️ GitHub Classroom fetch failed: {fetch_error}")
                await update.message.reply_text("\n".join(message_lines))
            else:
                entries.sort(key=lambda e: (e['deadline'] or datetime.max, e['name']))
                lines: List[str] = []
                for entry in entries:
                    lines.append(f"{entry['status']} \"{entry['name']}\"")
                    if entry['deadline']:
                        lines.append(f"Deadline: {entry['deadline'].strftime('%Y-%m-%d %H:%M:%S UTC')}{entry['time_remaining']}")
                    else:
                        lines.append("Deadline: N/A")
                    lines.append(f"Repository: {entry['repo_url'] or 'N/A'}")
                    if entry['teacher']:
                        lines.append(f"Teacher: {entry['teacher']}")
                    if entry['submission_status']:
                        lines.append(f"Submission status: {entry['submission_status']}")
                    if entry['submitted_at']:
                        lines.append(f"Submitted at: {entry['submitted_at']}")
                    lines.append(f"Source: {entry['source']}")
                    note_value = entry.get('note') or ''
                    note_value = note_value.strip()
                    if note_value:
                        lines.append("")
                        lines.append(f"Note: {note_value}")
                    lines.append("")

                output_text = "\n".join(lines).strip()
                if fetch_error:
                    output_text += f"\n\n⚠️ GitHub Classroom fetch failed: {fetch_error}"
                await update.message.reply_text(output_text)
        finally:
            db.close()

    async def add_assignment(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /add_assignment command - user provides repo link and deadline."""
        chat_id = update.effective_chat.id
        
        db = self.get_db()
        try:
            db_user = db.query(User).filter(User.telegram_id == chat_id).first()
            
            if not db_user:
                await update.message.reply_text("Please use /start first.")
                return
            
            if not db_user.github_token:
                await update.message.reply_text(
                    "Please register your GitHub token first:\n"
                    "/register_token <your_github_token>"
                )
                return

            if (db_user.role or 'student').lower() != 'teacher':
                await update.message.reply_text(
                    "Only teachers can add assignments. Use /set_role teacher <password> if you have teacher access."
                )
                return
            
            if not context.args or len(context.args) < 3:
                await update.message.reply_text(
                    "Usage: /add_assignment <name> <repo_link> <deadline>\n\n"
                    "Example: /add_assignment \"Homework 1\" https://github.com/org/repo \"Nov 11, 2025, 22:33 UTC\"\n\n"
                    "Or: /add_assignment \"Project\" org/repo-name \"Dec 31, 2024, 23:59 UTC\""
                )
                return
            
            name = context.args[0]
            repo_link = context.args[1]
            deadline_str = ' '.join(context.args[2:])
            
            # Remove quotes if present
            deadline_str = deadline_str.strip('"\'')
            
            try:
                # Parse date in format like "Nov 11, 2025, 22:33 UTC"
                deadline = date_parser.parse(deadline_str)
                # Ensure it's timezone-aware (UTC)
                if deadline.tzinfo is None:
                    deadline = deadline.replace(tzinfo=timezone.utc)
                # Convert to UTC naive datetime for storage
                deadline = deadline.astimezone(timezone.utc).replace(tzinfo=None)
            except (ValueError, TypeError) as e:
                await update.message.reply_text(
                    "Invalid deadline format. Use: \"Month Day, Year, HH:MM UTC\"\n"
                    "Example: \"Nov 11, 2025, 22:33 UTC\"\n"
                    "Or: \"Dec 31, 2024, 23:59 UTC\""
                )
                return
            
            # Parse repository name from link
            try:
                github_client = GitHubClient(token=db_user.github_token)
                repo_name = github_client.parse_repo_url(repo_link)
                
                if not repo_name:
                    await update.message.reply_text(
                        f"❌ Invalid repository link: {repo_link}\n"
                        f"Please provide a valid GitHub repository URL or org/repo format."
                    )
                    return
                
                # Validate repository exists using user's token
                repo_info = github_client.get_repository_activity(repo_name)
                
                if not repo_info['exists']:
                    await update.message.reply_text(
                        f"❌ Repository '{repo_name}' not found or not accessible.\n"
                        f"Please check the repository link and your token permissions."
                    )
                    return
                
                repo_url = repo_info.get('url', repo_link if repo_link.startswith('http') else f"https://github.com/{repo_name}")
            except Exception as e:
                await update.message.reply_text(
                    f"❌ Error accessing repository: {str(e)}\n"
                    f"Please check the repository link and your token permissions."
                )
                return
            
            # Create assignment
            assignment = Assignment(
                name=name,
                github_repo_name=repo_name,
                github_repo_url=repo_url,
                deadline=deadline,
                user_id=db_user.id
            )
            db.add(assignment)
            db.commit()
            
            await update.message.reply_text(
                f"✅ Assignment '{name}' added successfully!\n\n"
                f"Repository: {repo_name}\n"
                f"URL: {repo_url}\n"
                f"Deadline: {deadline.strftime('%Y-%m-%d %H:%M:%S UTC')}\n\n"
                f"The bot will monitor this assignment and notify you about the deadline."
            )
        finally:
            db.close()
    
    async def add_ci_repo(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /add_ci_repo command to start tracking CI status."""
        chat_id = update.effective_chat.id
        db = self.get_db()
        try:
            db_user = db.query(User).filter(User.telegram_id == chat_id).first()
            if not db_user:
                await update.message.reply_text("Please use /start first.")
                return
            if not db_user.github_token:
                await update.message.reply_text(
                    "Please register your GitHub token first:\n"
                    "/register_token <your_github_token>"
                )
                return
            if not context.args:
                await update.message.reply_text(
                    "Usage: /add_ci_repo <repo_url_or_owner/repo>\n\n"
                    "Example: /add_ci_repo https://github.com/org/project"
                )
                return

            raw_repo = ' '.join(context.args).strip('"\'')
            try:
                github_client = GitHubClient(token=db_user.github_token)
                repo_full_name = github_client.parse_repo_url(raw_repo)
                if not repo_full_name:
                    await update.message.reply_text(
                        f"❌ Не удалось распознать репозиторий: {raw_repo}\n"
                        f"Используйте формат https://github.com/owner/repo или owner/repo."
                    )
                    return

                repo_full_name = repo_full_name.strip().strip('/').lower()

                if not github_client.check_repository_exists(repo_full_name):
                    await update.message.reply_text(
                        f"❌ Репозиторий '{repo_full_name}' не найден или недоступен."
                    )
                    return
            except Exception as e:
                await update.message.reply_text(f"❌ Ошибка доступа к GitHub: {str(e)}")
                return

            existing = db.query(TrackedRepository).filter(
                TrackedRepository.user_id == db_user.id,
                TrackedRepository.repo_full_name == repo_full_name
            ).first()
            if existing:
                await update.message.reply_text(
                    f"ℹ️ Репозиторий '{repo_full_name}' уже отслеживается."
                )
                return

            repo_url = f"https://github.com/{repo_full_name}"
            tracked = TrackedRepository(
                user_id=db_user.id,
                repo_full_name=repo_full_name,
                repo_url=repo_url
            )
            db.add(tracked)
            db.commit()

            await update.message.reply_text(
                f"✅ Репозиторий '{repo_full_name}' добавлен для отслеживания CI."
            )
        finally:
            db.close()

    async def remove_ci_repo(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /remove_ci_repo command to stop tracking CI status."""
        chat_id = update.effective_chat.id
        db = self.get_db()
        try:
            db_user = db.query(User).filter(User.telegram_id == chat_id).first()
            if not db_user:
                await update.message.reply_text("Please use /start first.")
                return
            if not context.args:
                await update.message.reply_text(
                    "Usage: /remove_ci_repo <owner/repo>"
                )
                return

            raw_repo = ' '.join(context.args).strip('"\'')
            github_client = GitHubClient(token=db_user.github_token) if db_user.github_token else None
            repo_full_name = None
            if github_client:
                repo_full_name = github_client.parse_repo_url(raw_repo)
                if repo_full_name:
                    repo_full_name = repo_full_name.strip().strip('/').lower()
            if not repo_full_name:
                repo_full_name = raw_repo.strip().strip('/') if '/' in raw_repo else None
                if repo_full_name:
                    repo_full_name = repo_full_name.lower()
            if not repo_full_name:
                await update.message.reply_text(
                    f"❌ Не удалось распознать репозиторий: {raw_repo}"
                )
                return

            tracked = db.query(TrackedRepository).filter(
                TrackedRepository.user_id == db_user.id,
                TrackedRepository.repo_full_name == repo_full_name
            ).first()
            if not tracked:
                await update.message.reply_text(
                    f"ℹ️ Репозиторий '{repo_full_name}' не найден в списке отслеживания."
                )
                return

            db.delete(tracked)
            db.commit()
            await update.message.reply_text(
                f"✅ Репозиторий '{tracked.repo_full_name}' удалён из отслеживания."
            )
        finally:
            db.close()

    async def ci_status(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /ci_status command to show latest CI results."""
        chat_id = update.effective_chat.id
        db = self.get_db()
        try:
            db_user = db.query(User).filter(User.telegram_id == chat_id).first()
            if not db_user:
                await update.message.reply_text("Please use /start first.")
                return
            if not db_user.github_token:
                await update.message.reply_text(
                    "Please register your GitHub token first:\n"
                    "/register_token <your_github_token>"
                )
                return

            github_client = GitHubClient(token=db_user.github_token)

            repo_filter = None
            if context.args:
                filter_raw = ' '.join(context.args).strip('"\'')
                repo_filter = github_client.parse_repo_url(filter_raw)
                if not repo_filter:
                    await update.message.reply_text(
                        f"❌ Не удалось распознать репозиторий: {filter_raw}\n"
                        f"Используйте формат owner/repo."
                    )
                    return
                repo_filter = repo_filter.strip().strip('/').lower()

            query = db.query(TrackedRepository).filter(
                TrackedRepository.user_id == db_user.id
            )
            if repo_filter:
                query = query.filter(TrackedRepository.repo_full_name == repo_filter)

            tracked_repos = query.order_by(TrackedRepository.repo_full_name).all()
            if not tracked_repos:
                await update.message.reply_text(
                    "Список отслеживаемых репозиториев пуст.\n"
                    "Добавьте его командой /add_ci_repo <owner/repo>."
                )
                return

            responses = []
            for repo in tracked_repos:
                try:
                    status = github_client.get_ci_status(repo.repo_full_name)
                except Exception as e:
                    responses.append(
                        f"• {repo.repo_full_name}\n"
                        f"Не удалось получить статус CI: {str(e)}"
                    )
                    continue

                part = [f"• {repo.repo_full_name}"]
                part.append(status.get("message", ""))
                if status.get("html_url"):
                    part.append(f"Ссылка на запуск: {status['html_url']}")
                failure_summary = status.get("failure_summary")
                if failure_summary:
                    max_len = 1200
                    summary = failure_summary.strip()
                    if len(summary) > max_len:
                        summary = summary[:max_len] + "…"
                    part.append("Последние ошибки:")
                    part.append(summary)

                responses.append('\n'.join(filter(None, part)))

            await update.message.reply_text('\n\n'.join(responses))
        finally:
            db.close()
    
    async def classroom_assignments_overview(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /classroom_assignments to list assignments and participants."""
        chat_id = update.effective_chat.id
        db = self.get_db()
        try:
            db_user = db.query(User).filter(User.telegram_id == chat_id).first()
            if not db_user:
                await update.message.reply_text("Please use /start first.")
                return
            if (db_user.role or 'student').lower() != 'teacher':
                await update.message.reply_text(
                    "Only teachers can view classroom overviews. Use /set_role teacher <password> if you have teacher access."
                )
                return
            if not db_user.github_token:
                await update.message.reply_text(
                    "Please register your GitHub token first:\n"
                    "/register_token <your_github_token>"
                )
                return

            github_client = GitHubClient(token=db_user.github_token)

            classroom_filter = None
            if context.args:
                classroom_filter = ' '.join(context.args).strip('"\'').lower()

            try:
                classrooms = github_client.get_all_classrooms()
            except Exception as e:
                await update.message.reply_text(f"❌ Failed to load classrooms: {str(e)}")
                return

            matched_sections = []
            records_payload: List[Dict] = []

            for classroom in classrooms:
                class_name = classroom.get('name') or f"Classroom #{classroom.get('id')}"
                if classroom_filter and classroom_filter not in class_name.lower():
                    continue

                classroom_id = classroom.get('id')
                try:
                    assignments = github_client.get_assignments_for_classroom(classroom_id)
                except Exception as e:
                    matched_sections.append(
                        f"🏫 {class_name}\nНе удалось загрузить задания: {str(e)}"
                    )
                    continue

                if not assignments:
                    matched_sections.append(f"🏫 {class_name}\nЗаданий не найдено.")
                    continue

                section_lines = [f"🏫 {class_name}"]

                for assignment in assignments:
                    title = assignment.get('title') or 'Без названия'
                    raw_deadline = assignment.get('deadline')
                    deadline_dt = self._parse_datetime(raw_deadline)
                    if deadline_dt:
                        deadline_str = deadline_dt.strftime('%Y-%m-%d %H:%M UTC')
                    elif isinstance(raw_deadline, str):
                        deadline_str = raw_deadline
                    else:
                        deadline_str = 'N/A'

                    section_lines.append(f"📌 {title}")
                    section_lines.append(f"   Дедлайн: {deadline_str}")

                    try:
                        accepted = github_client.get_accepted_assignments(assignment, classroom_id)
                    except Exception as e:
                        section_lines.append(f"   Не удалось получить статус студентов: {str(e)}")
                        continue

                    try:
                        self._sync_assignment_record(db, db_user, classroom_id, class_name, assignment, accepted)
                    except Exception as sync_err:
                        print(f"Sync error for assignment {assignment.get('id')}: {sync_err}")

                    if not accepted:
                        section_lines.append("   Никто ещё не начал выполнение.")
                        records_payload.append({
                            'classroom_id': classroom_id,
                            'classroom_name': class_name,
                            'assignment_id': assignment.get('id'),
                            'assignment_title': title,
                            'deadline': deadline_dt,
                            'student_login': '',
                            'student_display_login': '',
                            'student_repo_url': '',
                            'submitted': None,
                            'passed': None,
                            'grade': None,
                            'commit_count': None,
                            'assignment_url': assignment.get('student_repository_url') or assignment.get('invitations_url') or '',
                            'raw': assignment,
                        })
                        continue

                    for acceptance in accepted:
                        login, repo_url, canonical_login = self._extract_student_identity(assignment, acceptance)
                        submitted = acceptance.get('submitted') or False
                        status_icon = "✅" if submitted else "⏳"
                        repo_text = f" – {repo_url}" if repo_url else ""
                        section_lines.append(f"   {status_icon} {login}{repo_text}")
                        records_payload.append({
                            'classroom_id': classroom_id,
                            'classroom_name': class_name,
                            'assignment_id': assignment.get('id'),
                            'assignment_title': title,
                            'deadline': deadline_dt,
                            'student_login': canonical_login or login,
                            'student_display_login': login,
                            'student_repo_url': repo_url,
                            'submitted': bool(submitted),
                            'passed': acceptance.get('passed'),
                            'grade': acceptance.get('grade'),
                            'commit_count': acceptance.get('commit_count'),
                            'assignment_url': repo_url or assignment.get('student_repository_url') or assignment.get('invitations_url') or '',
                            'raw': acceptance,
                        })

                matched_sections.append('\n'.join(section_lines))

            try:
                self._store_classroom_records(db, db_user, records_payload)
            except Exception as store_err:
                print(f"Failed to store classroom snapshot: {store_err}")

            if not matched_sections:
                await update.message.reply_text(
                    "Классы не найдены. Уточните название: /classroom_assignments <часть названия>"
                )
                return

            max_len = 3500
            current_chunk = ""
            chunks = []

            for section in matched_sections:
                section = section.strip()
                if not section:
                    continue
                addition = section if not current_chunk else f"{current_chunk}\n\n{section}"
                if len(addition) > max_len and current_chunk:
                    chunks.append(current_chunk)
                    if len(section) > max_len:
                        # Split overly long section
                        start = 0
                        while start < len(section):
                            chunks.append(section[start:start + max_len])
                            start += max_len
                        current_chunk = ""
                    else:
                        current_chunk = section
                else:
                    current_chunk = addition

            if current_chunk:
                chunks.append(current_chunk)

            for chunk in chunks:
                await update.message.reply_text(chunk)
        finally:
            db.close()
    
    async def export_classroom_excel(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /export_assignments_excel to export classroom data to Excel."""
        chat_id = update.effective_chat.id
        db = self.get_db()
        try:
            db_user = db.query(User).filter(User.telegram_id == chat_id).first()
            
            if not db_user:
                await update.message.reply_text("Please use /start first.")
                return
            
            if (db_user.role or 'student').lower() != 'teacher':
                await update.message.reply_text(
                    "Only teachers can export classroom data. Use /set_role teacher <password> if you have teacher access."
                )
                return
            
            if not db_user.github_token:
                await update.message.reply_text(
                    "Please register your GitHub token first:\n"
                    "/register_token <your_github_token>"
                )
                return

            github_client = GitHubClient(token=db_user.github_token)

            classroom_filter = None
            if context.args:
                classroom_filter = ' '.join(context.args).strip('"\'').lower()

            try:
                classrooms = github_client.get_all_classrooms()
            except Exception as e:
                await update.message.reply_text(f"❌ Failed to load classrooms: {str(e)}")
                return

            detail_rows = []
            summary_rows = []

            for classroom in classrooms:
                class_name = classroom.get('name') or f"Classroom #{classroom.get('id')}"
                if classroom_filter and classroom_filter not in class_name.lower():
                    continue

                classroom_id = classroom.get('id')
                try:
                    assignments = github_client.get_assignments_for_classroom(classroom_id)
                except Exception as e:
                    summary_rows.append({
                        'class': class_name,
                        'assignment': '—',
                        'started': 0,
                        'submitted': 0,
                        'deadline': '',
                        'error': str(e),
                    })
                    continue

                if not assignments:
                    summary_rows.append({
                        'class': class_name,
                        'assignment': '—',
                        'started': 0,
                        'submitted': 0,
                        'deadline': '',
                        'error': '',
                    })
                    continue

                for assignment in assignments:
                    title = assignment.get('title') or 'Untitled'
                    deadline = assignment.get('deadline')
                    if isinstance(deadline, datetime):
                        deadline_obj = deadline
                    else:
                        deadline_obj = None
                        if isinstance(deadline, str):
                            try:
                                deadline_obj = date_parser.parse(deadline)
                            except Exception:
                                deadline_obj = None
                    if deadline_obj:
                        if deadline_obj.tzinfo:
                            deadline_obj = deadline_obj.astimezone(timezone.utc)
                        deadline_str = deadline_obj.strftime('%Y-%m-%d %H:%M:%S UTC')
                    elif isinstance(deadline, str):
                        deadline_str = deadline
                    else:
                        deadline_str = ''

                    try:
                        accepted = github_client.get_accepted_assignments(assignment, classroom_id)
                    except Exception as e:
                        summary_rows.append({
                            'class': class_name,
                            'assignment': title,
                            'started': 0,
                            'submitted': 0,
                            'deadline': deadline_str,
                            'error': str(e),
                        })
                        continue

                    try:
                        self._sync_assignment_record(db, db_user, classroom_id, class_name, assignment, accepted)
                    except Exception as sync_err:
                        print(f"Sync error during export for assignment {assignment.get('id')}: {sync_err}")

                    started_count = len(accepted)
                    submitted_count = sum(1 for acc in accepted if acc.get('submitted'))

                    if not accepted:
                        detail_rows.append([
                            class_name,
                            title,
                            assignment.get('id'),
                            deadline_str,
                            "—",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                        ])
                    else:
                        for acceptance in accepted:
                            login, repo_url, _ = self._extract_student_identity(assignment, acceptance)
                            submitted = acceptance.get('submitted') or False
                            passed = acceptance.get('passed') or False
                            grade = acceptance.get('grade')
                            commit_count = acceptance.get('commit_count')
                            updated_at = acceptance.get('updated_at') or acceptance.get('created_at')
                            if isinstance(updated_at, str):
                                updated_str = updated_at
                            elif isinstance(updated_at, datetime):
                                updated_obj = updated_at.astimezone(timezone.utc) if updated_at.tzinfo else updated_at
                                updated_str = updated_obj.strftime('%Y-%m-%d %H:%M:%S UTC')
                            else:
                                updated_str = ''

                            detail_rows.append([
                                class_name,
                                title,
                                assignment.get('id'),
                                deadline_str,
                                login,
                                repo_url,
                                "Да" if submitted else "Нет",
                                "Да" if passed else "Нет",
                                grade if grade is not None else "",
                                commit_count if commit_count is not None else "",
                                updated_str,
                            ])

                    summary_rows.append({
                        'class': class_name,
                        'assignment': title,
                        'started': started_count,
                        'submitted': submitted_count,
                        'deadline': deadline_str,
                        'error': '',
                    })

            if not detail_rows and not summary_rows:
                await update.message.reply_text(
                    "Нет данных для экспорта. Убедитесь, что вы указали верный фильтр или что в классах есть задания."
                )
                return

            wb = Workbook()
            ws_details = wb.active
            ws_details.title = "Students"
            detail_header = [
                "Classroom",
                "Assignment",
                "Assignment ID",
                "Deadline (UTC)",
                "Student",
                "Repository URL",
                "Submitted",
                "Passed",
                "Grade",
                "Commit Count",
                "Last Update",
            ]
            ws_details.append(detail_header)
            for row in detail_rows:
                ws_details.append(row)

            for idx, column_cells in enumerate(ws_details.columns, start=1):
                max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
                ws_details.column_dimensions[get_column_letter(idx)].width = min(max_length + 2, 60)

            ws_summary = wb.create_sheet("Summary")
            summary_header = ["Classroom", "Assignment", "Students Started", "Submitted", "Deadline (UTC)", "Error"]
            ws_summary.append(summary_header)
            for entry in summary_rows:
                ws_summary.append([
                    entry['class'],
                    entry['assignment'],
                    entry['started'],
                    entry['submitted'],
                    entry['deadline'],
                    entry['error'],
                ])

            for idx, column_cells in enumerate(ws_summary.columns, start=1):
                max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
                ws_summary.column_dimensions[get_column_letter(idx)].width = min(max_length + 2, 50)

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            filename = f"classroom_export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
            caption_parts = ["Экспорт завершён."]
            if detail_rows:
                caption_parts.append(f"Строк в деталях: {len(detail_rows)}.")
            if summary_rows:
                caption_parts.append(f"Записей в сводке: {len(summary_rows)}.")
            caption = " ".join(caption_parts)

            await update.message.reply_document(
                document=buffer,
                filename=filename,
                caption=caption
            )
        finally:
            db.close()
    
    async def export_users_excel(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /export_users_excel to export user information."""
        chat_id = update.effective_chat.id
        db = self.get_db()
        try:
            db_user = db.query(User).filter(User.telegram_id == chat_id).first()
            
            if not db_user:
                await update.message.reply_text("Please use /start first.")
                return
            
            if (db_user.role or 'student').lower() != 'teacher':
                await update.message.reply_text(
                    "Only teachers can export user information. Use /set_role teacher <password> if you have teacher access."
                )
                return

            users = (
                db.query(User)
                .options(
                    joinedload(User.assignments),
                    joinedload(User.submissions),
                    joinedload(User.ci_repositories)
                )
                .order_by(User.created_at)
                .all()
            )

            if not users:
                await update.message.reply_text("No users found in the database.")
                return

            wb = Workbook()
            ws_users = wb.active
            ws_users.title = "Users"
            header = [
                "User ID",
                "Telegram ID",
                "Telegram Username",
                "First Name",
                "Last Name",
                "Role",
                "GitHub Username",
                "Has GitHub Token",
                "Assignments Owned",
                "Submissions Linked",
                "Submissions Submitted",
                "Tracked Repositories",
                "Notification Threshold (h)",
                "Notification Period (s)",
                "Created At (UTC)",
                "Last Activity (UTC)",
            ]
            ws_users.append(header)

            role_counter = Counter()
            token_counter = Counter()
            submission_total = 0
            submission_submitted_total = 0
            assignments_total = 0

            def format_datetime(dt: Optional[datetime]) -> str:
                if not dt:
                    return ""
                if dt.tzinfo:
                    dt_local = dt.astimezone(timezone.utc)
                else:
                    dt_local = dt
                return dt_local.strftime('%Y-%m-%d %H:%M:%S UTC')

            for user in users:
                role_name = (user.role or 'student').lower()
                role_counter[role_name] += 1
                has_token = bool(user.github_token)
                token_counter['with_token' if has_token else 'without_token'] += 1

                owned_assignments = len(user.assignments or [])
                assignments_total += owned_assignments
                submissions = user.submissions or []
                submission_count = len(submissions)
                submission_total += submission_count
                submitted_count = sum(1 for s in submissions if s.is_submitted)
                submission_submitted_total += submitted_count
                tracked_repos = len(user.ci_repositories or [])

                notify_threshold = user.notify_threshold_hours if user.notify_threshold_hours is not None else ""
                notify_period = user.notify_period_seconds if user.notify_period_seconds is not None else ""

                activity_candidates: List[datetime] = []
                if user.created_at:
                    activity_candidates.append(user.created_at)
                for assignment in user.assignments or []:
                    if assignment.updated_at:
                        activity_candidates.append(assignment.updated_at)
                    if assignment.created_at:
                        activity_candidates.append(assignment.created_at)
                for submission in submissions:
                    for attr in (submission.updated_at, submission.submitted_at, submission.created_at):
                        if attr:
                            activity_candidates.append(attr)
                last_activity = max(activity_candidates) if activity_candidates else None

                ws_users.append([
                    user.id,
                    user.telegram_id,
                    user.username or "",
                    user.first_name or "",
                    user.last_name or "",
                    role_name.title(),
                    user.github_username or "",
                    "Yes" if has_token else "No",
                    owned_assignments,
                    submission_count,
                    submitted_count,
                    tracked_repos,
                    notify_threshold,
                    notify_period,
                    format_datetime(user.created_at),
                    format_datetime(last_activity),
                ])

            for idx, column_cells in enumerate(ws_users.columns, start=1):
                max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
                ws_users.column_dimensions[get_column_letter(idx)].width = min(max_length + 2, 50)

            ws_summary = wb.create_sheet("Summary")
            ws_summary.append(["Metric", "Value"])
            ws_summary.append(["Total users", len(users)])
            ws_summary.append(["Users with GitHub token", token_counter.get('with_token', 0)])
            ws_summary.append(["Users without GitHub token", token_counter.get('without_token', 0)])
            ws_summary.append(["Total assignments owned", assignments_total])
            ws_summary.append(["Total submissions linked", submission_total])
            ws_summary.append(["Total submissions submitted", submission_submitted_total])

            ws_summary.append([])
            ws_summary.append(["Users by role", "Count"])
            for role_name, count in role_counter.most_common():
                ws_summary.append([role_name.title(), count])

            for idx, column_cells in enumerate(ws_summary.columns, start=1):
                max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
                ws_summary.column_dimensions[get_column_letter(idx)].width = min(max_length + 2, 40)

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            filename = f"users_export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
            caption = f"Экспортировано пользователей: {len(users)}."

            await update.message.reply_document(
                document=buffer,
                filename=filename,
                caption=caption
            )
        finally:
            db.close()

    async def dump_submissions(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /dump_submissions command to list all submission records."""
        chat_id = update.effective_chat.id
        db = self.get_db()
        try:
            db_user = db.query(User).filter(User.telegram_id == chat_id).first()

            if not db_user:
                await update.message.reply_text("Please use /start first.")
                return

            if (db_user.role or 'student').lower() != 'teacher':
                await update.message.reply_text(
                    "Only teachers can dump assignment data. Use /set_role teacher <password> if you have teacher access."
                )
                return

            submissions = (
                db.query(Submission)
                .options(joinedload(Submission.assignment).joinedload(Assignment.user),
                         joinedload(Submission.user))
                .order_by(Submission.id)
                .all()
            )

            if not submissions:
                await update.message.reply_text("Submission table is empty.")
                return

            def format_value(value):
                if value is None:
                    return "NULL"
                if isinstance(value, datetime):
                    dt = value if value.tzinfo is None else value.astimezone(timezone.utc)
                    return dt.strftime('%Y-%m-%d %H:%M:%S UTC')
                return str(value)

            lines: List[str] = []
            for submission in submissions:
                assignment = submission.assignment
                assignment_info = ""
                if assignment:
                    owner = assignment.user
                    owner_info = ""
                    if owner:
                        owner_info = owner.github_username or owner.username or owner.first_name or ""
                    assignment_info = (
                        f"{assignment.id} | {assignment.name} | owner={owner_info}"
                    )
                student = submission.user
                student_info = ""
                if student:
                    student_info = student.github_username or student.username or student.first_name or ""
                lines.append("────────────────────────")
                lines.append(f"id: {submission.id}")
                lines.append(f"assignment_id: {submission.assignment_id}")
                lines.append(f"user_id: {submission.user_id}")
                lines.append(f"user_login: {student_info}")
                lines.append(f"assignment_info: {assignment_info}")
                lines.append(f"github_repo_url: {submission.github_repo_url or ''}")
                lines.append(f"last_commit_sha: {submission.last_commit_sha or ''}")
                lines.append(f"last_commit_date: {format_value(submission.last_commit_date)}")
                lines.append(f"is_submitted: {submission.is_submitted}")
                lines.append(f"submitted_at: {format_value(submission.submitted_at)}")
                lines.append(f"created_at: {format_value(submission.created_at)}")
                lines.append(f"updated_at: {format_value(submission.updated_at)}")
                lines.append("")

            chunk_size = 3500
            current = []
            current_length = 0

            for line in lines:
                add_len = len(line) + 1
                if current_length + add_len > chunk_size:
                    await update.message.reply_text("\n".join(current).strip())
                    current = []
                    current_length = 0
                current.append(line)
                current_length += add_len

            if current:
                await update.message.reply_text("\n".join(current).strip())
        finally:
            db.close()
    
    async def delete_assignment(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /delete_assignment command - delete an assignment by name."""
        chat_id = update.effective_chat.id
        
        db = self.get_db()
        try:
            db_user = db.query(User).filter(User.telegram_id == chat_id).first()
            
            if not db_user:
                await update.message.reply_text("Please use /start first.")
                return
            
            if (db_user.role or 'student').lower() != 'teacher':
                await update.message.reply_text(
                    "Only teachers can delete assignments. Use /set_role teacher <password> if you have teacher access."
                )
                return
            
            if not context.args:
                await update.message.reply_text(
                    "Usage: /delete_assignment <assignment_name>\n\n"
                    "Example: /delete_assignment \"Homework 1\"\n\n"
                    "Use /assignments to see your assignments."
                )
                return
            
            assignment_name = ' '.join(context.args)
            
            # Find assignment by name (case-insensitive) that belongs to this user
            assignment = db.query(Assignment).filter(
                and_(
                    Assignment.user_id == db_user.id,
                    Assignment.name.ilike(f"%{assignment_name}%")
                )
            ).first()
            
            if not assignment:
                await update.message.reply_text(
                    f"❌ Assignment '{assignment_name}' not found.\n\n"
                    f"Use /assignments to see your assignments."
                )
                return
            
            # Store assignment name for confirmation message
            deleted_name = assignment.name
            
            # Delete the assignment (cascade will handle related submissions)
            db.delete(assignment)
            db.commit()
            
            await update.message.reply_text(
                f"✅ Assignment '{deleted_name}' deleted successfully!"
            )
        finally:
            db.close()

    async def set_my_notify_threshold(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /set_my_notify_threshold <days> to control when your reminders start."""
        chat_id = update.effective_chat.id
        db = self.get_db()
        try:
            if not context.args:
                await update.message.reply_text(
                    "Usage: /set_my_notify_threshold <days>\n"
                    "Example: /set_my_notify_threshold 7"
                )
                return
            try:
                days = int(context.args[0])
                if days < 0:
                    raise ValueError()
            except ValueError:
                await update.message.reply_text("Days must be a non-negative integer.")
                return
            db_user = db.query(User).filter(User.telegram_id == chat_id).first()
            if not db_user:
                await update.message.reply_text("Please use /start first.")
                return
            db_user.notify_threshold_hours = days * 24
            db.commit()
            await update.message.reply_text(
                f"✅ Your notification threshold set to {days} day(s) before deadline."
            )
        finally:
            db.close()

    async def set_my_notify_period(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /set_my_notify_period <value><m|h> to control your reminder frequency."""
        chat_id = update.effective_chat.id
        db = self.get_db()
        try:
            if not context.args:
                await update.message.reply_text(
                    "Usage: /set_my_notify_period <value><m|h>\n"
                    "Examples: /set_my_notify_period 60m, /set_my_notify_period 1h"
                )
                return
            token = context.args[0].strip().lower()
            import re
            m = re.match(r"^(\d+)(m|h)$", token)
            if not m:
                await update.message.reply_text(
                    "Invalid format. Use <value><m|h>, e.g. 30m or 2h."
                )
                return
            value = int(m.group(1))
            unit = m.group(2)
            seconds = value * 60 if unit == 'm' else value * 3600
            db_user = db.query(User).filter(User.telegram_id == chat_id).first()
            if not db_user:
                await update.message.reply_text("Please use /start first.")
                return
            db_user.notify_period_seconds = seconds
            db.commit()
            await update.message.reply_text(
                f"✅ Your notification period set to {value}{unit}."
            )
        finally:
            db.close()

    async def set_role(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /set_role <student|teacher> [password]."""
        chat_id = update.effective_chat.id
        db = self.get_db()
        try:
            if not context.args:
                await update.message.reply_text(
                    "Usage: /set_role <student|teacher> [password]\n"
                    "Teachers must supply the shared access password."
                )
                return

            desired_role = context.args[0].strip().lower()
            if desired_role not in ('student', 'teacher'):
                await update.message.reply_text("Role must be either 'student' or 'teacher'.")
                return

            db_user = db.query(User).filter(User.telegram_id == chat_id).first()
            if not db_user:
                await update.message.reply_text("Please use /start first.")
                return

            if desired_role == 'teacher':
                if not self.teacher_password:
                    await update.message.reply_text(
                        "Teacher role is not configured by the administrator."
                    )
                    return
                if len(context.args) < 2:
                    await update.message.reply_text("Teacher role requires a password.")
                    return
                provided_password = context.args[1]
                if provided_password != self.teacher_password:
                    await update.message.reply_text("Invalid teacher password.")
                    return

            db_user.role = desired_role
            db.commit()
            await update.message.reply_text(
                f"✅ Your role has been updated to '{desired_role.title()}'."
            )
        finally:
            db.close()

    async def add_note(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /add_note <assignment_name> <text> to annotate an assignment."""
        chat_id = update.effective_chat.id
        db = self.get_db()
        try:
            if not context.args or len(context.args) < 2:
                await update.message.reply_text(
                    "Usage: /add_note <assignment_name> <text>\n"
                    "Example: /add_note data-test-2 Students CI has out of memory error"
                )
                return

            db_user = db.query(User).filter(User.telegram_id == chat_id).first()
            if not db_user:
                await update.message.reply_text("Please use /start first.")
                return

            if (db_user.role or 'student').lower() != 'teacher':
                await update.message.reply_text(
                    "Only teachers can add notes. Use /set_role teacher <password> if you have teacher access."
                )
                return

            assignment_name = context.args[0]
            note_text = ' '.join(context.args[1:]).strip()
            if not note_text:
                await update.message.reply_text("Note text cannot be empty.")
                return

            assignment = db.query(Assignment).filter(
                and_(
                    Assignment.user_id == db_user.id,
                    Assignment.name.ilike(f"%{assignment_name}%")
                )
            ).first()
            if not assignment:
                await update.message.reply_text(
                    f"❌ Assignment '{assignment_name}' not found.\n\n"
                    "Use /assignments to see your assignments."
                )
                return

            assignment.note = note_text
            db.commit()
            await update.message.reply_text(
                f"✅ Note for assignment '{assignment.name}' added successfully!"
            )
        finally:
            db.close()

    async def delete_note(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /delete_note <assignment_name> to remove a note."""
        chat_id = update.effective_chat.id
        db = self.get_db()
        try:
            if not context.args:
                await update.message.reply_text(
                    "Usage: /delete_note <assignment_name>\n"
                    "Examples: /delete_note my-homework"
                )
                return

            db_user = db.query(User).filter(User.telegram_id == chat_id).first()
            if not db_user:
                await update.message.reply_text("Please use /start first.")
                return

            if (db_user.role or 'student').lower() != 'teacher':
                await update.message.reply_text(
                    "Only teachers can delete notes. Use /set_role teacher <password> if you have teacher access."
                )
                return

            assignment_name = context.args[0]
            assignment = db.query(Assignment).filter(
                and_(
                    Assignment.user_id == db_user.id,
                    Assignment.name.ilike(f"%{assignment_name}%")
                )
            ).first()
            if not assignment:
                await update.message.reply_text(
                    f"❌ Assignment '{assignment_name}' not found.\n\n"
                    f"Use /assignments to see your assignments."
                )
                return
            assignment.note = ''
            db.commit()
            await update.message.reply_text(
                f"✅ Note for assignment '{assignment.name}' deleted successfully!"
            )
        finally:
            db.close()

    async def set_github_username(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /set_github_username <username>."""
        chat_id = update.effective_chat.id
        db = self.get_db()
        try:
            if not context.args:
                await update.message.reply_text(
                    "Usage: /set_github_username <username>"
                )
                return

            github_username = context.args[0].strip()
            if not re.match(r"^[A-Za-z0-9-]{1,39}$", github_username):
                await update.message.reply_text(
                    "Invalid GitHub username. It should be 1-39 characters, containing letters, numbers, or hyphens."
                )
                return

            db_user = db.query(User).filter(User.telegram_id == chat_id).first()
            if not db_user:
                await update.message.reply_text("Please use /start first.")
                return

            db_user.github_username = github_username
            db.commit()
            await update.message.reply_text(
                f"✅ GitHub username set to {github_username}."
            )
        finally:
            db.close()

def main():
    """Main function to run the bot."""
    # Validate configuration
    try:
        Config.validate()
    except ValueError as e:
        print(f"Configuration error: {e}")
        return
    
    # Initialize database
    init_db()
    
    # Create bot application
    application = Application.builder().token(Config.TELEGRAM_BOT_TOKEN).build()
    
    # Create bot instance
    bot_instance = HomeworkTrackerBot()
    
    # Add handlers
    application.add_handler(CommandHandler("start", bot_instance.start))
    application.add_handler(CommandHandler("help", bot_instance.help_command))
    application.add_handler(CommandHandler("register_token", bot_instance.register_token))
    application.add_handler(CommandHandler("assignments", bot_instance.list_assignments))
    application.add_handler(CommandHandler("add_assignment", bot_instance.add_assignment))
    application.add_handler(CommandHandler("add_ci_repo", bot_instance.add_ci_repo))
    application.add_handler(CommandHandler("remove_ci_repo", bot_instance.remove_ci_repo))
    application.add_handler(CommandHandler("ci_status", bot_instance.ci_status))
    application.add_handler(CommandHandler("classroom_assignments", bot_instance.classroom_assignments_overview))
    application.add_handler(CommandHandler("export_assignments_excel", bot_instance.export_classroom_excel))
    application.add_handler(CommandHandler("export_users_excel", bot_instance.export_users_excel))
    application.add_handler(CommandHandler("dump_submissions", bot_instance.dump_submissions))
    application.add_handler(CommandHandler("delete_assignment", bot_instance.delete_assignment))
    application.add_handler(CommandHandler("set_my_notify_threshold", bot_instance.set_my_notify_threshold))
    application.add_handler(CommandHandler("set_my_notify_period", bot_instance.set_my_notify_period))
    application.add_handler(CommandHandler("set_role", bot_instance.set_role))
    application.add_handler(CommandHandler("set_github_username", bot_instance.set_github_username))
    application.add_handler(CommandHandler("add_note", bot_instance.add_note))
    application.add_handler(CommandHandler("delete_note", bot_instance.delete_note))
    
    # Start the bot
    print("Bot is starting...")
    application.run_polling(allowed_updates=Update.ALL_TYPES)

if __name__ == '__main__':
    main()
